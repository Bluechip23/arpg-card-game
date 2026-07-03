#!/usr/bin/env python3
"""Builds /tmp/export/arpg_status_effects.xlsx (Buffs + Debuffs tabs) by parsing
scripts/effects/buff.gd and scripts/effects/debuff.gd directly — no Godot
needed. Google Sheets imports the .xlsx as-is.

Each row: Name | Enum ID | Category | Effect | Icon Hex | colour swatch cell.
The icon colours are the ones BuffIconUI/DebuffIconUI already use, so the
sheet doubles as the badge-design reference.

Usage:  python3 tests/build_status_effects_spreadsheet.py
"""
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = "/tmp/export/arpg_status_effects.xlsx"

CATEGORIES = {
    # Buffs
    "STRENGTHEN": "Offense", "ENLIGHTENED": "Offense", "LIFE_STEAL": "Offense",
    "WEAR_DOWN": "Offense", "ARMOR_BREAK": "Offense",
    "FORTIFY": "Defense", "BOLSTER": "Defense", "SMITH": "Defense",
    "BRACE": "Defense", "RESILIENT": "Defense", "THORNS": "Defense",
    "SHIELD_READY": "Defense", "REPELLED_BLOCK": "Defense", "SHIELD_OF_GROWTH": "Defense",
    "REGEN": "Sustain", "MORPHINE": "Sustain", "PHOENIX_GRACE": "Sustain", "CLEANSE": "Sustain",
    "FOCUSED": "Economy", "BLESSED": "Economy", "STEADY": "Economy",
    "HASTE": "Economy", "DEMONIC_RAGE": "Economy", "INVISIBLE": "Economy",
    # Debuffs
    "BLEED": "Damage over time", "BURN": "Damage over time", "POISON": "Damage over time",
    "SHOCKED": "Damage over time", "CURSED": "Damage over time", "DRAIN": "Damage over time",
    "STUN": "Action lockout", "FROZEN": "Action lockout", "DISARM": "Action lockout",
    "SILENCE": "Action lockout", "CUFFED": "Action lockout",
    "ROOTED": "Movement", "SLOWED": "Movement", "INEBRIATE": "Movement",
    "TETHERED": "Movement", "MAGNETIZED": "Movement",
    "HEXED": "Card disruption", "LOCKED": "Card disruption", "WEIGHTED": "Card disruption",
    "STAGGERED": "Card disruption", "CLUMSY": "Card disruption",
    "VULNERABLE": "Defense shred", "EXPOSED": "Defense shred", "BRITTLE": "Defense shred",
    "COLD": "Defense shred", "BLIND": "Defense shred", "LINKED": "Defense shred",
}


def parse_effects(path: str, prefix: str):
    """Yield (enum_id, display_name, description, hex_color) from a buff/debuff .gd file."""
    src = open(path).read()
    names, descs, colors = {}, {}, {}
    # match arms in the _init/setup: PREFIX.ID: ... name = "..." ... description = "..."
    arm = None
    for line in src.splitlines():
        m = re.match(rf"\s*{prefix}\.(\w+)\s*:\s*(#.*)?$", line)
        if m:
            arm = m.group(1)
            continue
        m = re.match(r'\s*(?:buff|debuff)_name\s*=\s*"([^"]+)"', line)
        if m and arm:
            names[arm] = m.group(1)
            continue
        m = re.match(r'\s*description\s*=\s*"([^"]+)"', line)
        if m and arm and arm not in descs:
            descs[arm] = m.group(1)
            continue
    # icon colours: PREFIX.ID: return Color(r, g, b)
    for m in re.finditer(
        rf"{prefix}\.(\w+):\s*return\s*Color\(([\d.]+),\s*([\d.]+),\s*([\d.]+)\)", src
    ):
        rgb = tuple(int(round(float(c) * 255)) for c in m.groups()[1:])
        colors[m.group(1)] = "%02X%02X%02X" % rgb
    for enum_id, name in names.items():
        desc = descs.get(enum_id, "")
        # Turn format placeholders into a readable "X": "%d damage" -> "X damage"
        desc = desc.replace("%%", "\x00").replace("%d", "X").replace("\x00", "%")
        yield enum_id, name, desc, colors.get(enum_id, "FFFFFF")


def write_sheet(ws, rows):
    header = ["Name", "Enum ID", "Category", "Effect", "Icon Hex", "Swatch"]
    ws.append(header)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="2A2A35")
        cell.font = Font(bold=True, color="FFFFFF")
    for enum_id, name, desc, hexcol in rows:
        ws.append([name, enum_id, CATEGORIES.get(enum_id, ""), desc, "#" + hexcol, ""])
        ws.cell(ws.max_row, 6).fill = PatternFill("solid", fgColor=hexcol)
    widths = [18, 20, 18, 72, 10, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    for row in ws.iter_rows(min_row=2):
        row[3].alignment = Alignment(wrap_text=True)
    ws.freeze_panes = "A2"


def sort_rows(rows):
    order = list(dict.fromkeys(CATEGORIES.values()))
    return sorted(rows, key=lambda r: (order.index(CATEGORIES.get(r[0], "")), r[1]))


buffs = sort_rows(list(parse_effects(f"{ROOT}/scripts/effects/buff.gd", "BuffType")))
debuffs = sort_rows(list(parse_effects(f"{ROOT}/scripts/effects/debuff.gd", "DebuffType")))

wb = Workbook()
write_sheet(wb.active, buffs)
wb.active.title = "Buffs"
write_sheet(wb.create_sheet("Debuffs"), debuffs)

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f"Wrote {OUT}: {len(buffs)} buffs, {len(debuffs)} debuffs")
